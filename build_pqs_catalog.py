import json, os, re
from openpyxl import load_workbook

BASE_DIR = "/mnt/data"
PQS_FILE = os.path.join(BASE_DIR, "PQS Rev Z(1).xlsm")
OPT_FILE = os.path.join(BASE_DIR, "Machine OPTs(1).xlsm")
STD_FILE = os.path.join(BASE_DIR, "Machine STDs(1).xlsm")
CNC_FILE = os.path.join(BASE_DIR, "CNC Controls(1).xlsm")
OUT_FILE = os.path.join(BASE_DIR, "pqs_catalog_live.json")

code_pat = re.compile(r'^(?:ST|OP|F)\d{3,4}(?:-\d{2,3})?[A-Z0-9-]*$')

def parse_option_sheet(ws, section):
    base_price = None
    lines = []
    current = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        h = ws.cell(r, 8).value

        if isinstance(a, str) and 'base' in a.lower() and 'u.s. dollars' in a.lower():
            if isinstance(h, (int, float)):
                base_price = float(h)
            elif isinstance(h, str):
                try:
                    base_price = float(h.replace(',', '').replace('$', ''))
                except ValueError:
                    pass
            continue

        code = a.strip() if isinstance(a, str) else None
        if code and code_pat.match(code):
            if current:
                lines.append(current)

            standard = isinstance(h, str) and h.strip().upper() == 'STD'
            price = None
            if standard:
                price = 0.0
            elif isinstance(h, (int, float)):
                price = float(h)
            elif isinstance(h, str):
                txt = h.strip().upper()
                if txt != 'TBD':
                    try:
                        price = float(txt.replace(',', '').replace('$', ''))
                    except ValueError:
                        price = None

            current = {
                'code': code,
                'description': (b or '').strip() if isinstance(b, str) else (str(b).strip() if b is not None else ''),
                'detailLines': [],
                'section': section,
                'standard': standard,
                'unitPrice': price,
                'priceStatus': 'STD' if standard else ('TBD' if price is None else 'PRICED'),
                'sourceSheet': ws.title,
                'sourceRow': r,
            }
        else:
            if current and b is not None and str(b).strip():
                current['detailLines'].append(str(b).strip())

    if current:
        lines.append(current)
    return base_price, lines

def parse_std_specs(ws):
    sections = []
    current = None
    for r in range(1, ws.max_row + 1):
        b = ws.cell(r, 2).value
        c = ws.cell(r, 3).value
        if isinstance(b, str) and re.match(r'^\d+\)$', b.strip()) and c:
            if current:
                sections.append(current)
            current = {'title': str(c).strip(), 'lines': [], 'sourceRow': r}
        elif current and c is not None and str(c).strip():
            current['lines'].append(str(c).strip())
    if current:
        sections.append(current)
    return sections

def infer_group(code, desc):
    text = f"{code} {desc}".lower()
    if 'spindle' in text:
        return 'spindle_family'
    if 'tool changer' in text or 'atc' in text:
        return 'atc_family'
    if 'chip conveyor' in text or 'conveyor' in text:
        return 'chip_conveyor'
    if 'probe' in text:
        return 'probe_system'
    if 'coolant' in text:
        return 'coolant'
    return None

def main():
    pqs_wb = load_workbook(PQS_FILE, data_only=False)
    opt_wb = load_workbook(OPT_FILE, data_only=False)
    std_wb = load_workbook(STD_FILE, data_only=False)
    cnc_wb = load_workbook(CNC_FILE, data_only=False)

    info_ws = pqs_wb['Info']
    active_models = []
    for r in range(34, 90):
        idx = info_ws.cell(r, 1).value
        model = info_ws.cell(r, 2).value
        if isinstance(idx, (int, float)) and idx <= 56 and isinstance(model, str) and model.strip():
            active_models.append(model.strip())

    catalog = {
        'companyProfile': {
            'brandName': 'Mitsui Seiki U.S.A., Inc.',
            'sourceFiles': [os.path.basename(PQS_FILE), os.path.basename(OPT_FILE), os.path.basename(STD_FILE), os.path.basename(CNC_FILE)],
            'notes': 'Generated from uploaded source workbooks. Active machine list driven by PQS Rev Z Info tab.'
        },
        'machines': []
    }

    for model in active_models:
        base_price, machine_lines = parse_option_sheet(opt_wb[model], 'machine')
        _, cnc_lines = parse_option_sheet(cnc_wb[model], 'cnc')
        std_sections = parse_std_specs(std_wb[model])

        for line in machine_lines + cnc_lines:
            line['group'] = infer_group(line['code'], line['description'])
            line['fullDescription'] = line['description'] + ((' — ' + ' '.join(line['detailLines'])) if line['detailLines'] else '')

        catalog['machines'].append({
            'id': model,
            'label': model,
            'basePrice': base_price,
            'machineOptions': machine_lines,
            'cncOptions': cnc_lines,
            'standardSpecSections': std_sections
        })

    with open(OUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, indent=2)

    print(f"Wrote {OUT_FILE} with {len(catalog['machines'])} active machines.")

if __name__ == "__main__":
    main()
